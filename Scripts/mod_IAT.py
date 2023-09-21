import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import coordinate_to_tuple

#csv_folder = r"C:\Users\Christian\RTF\EULR\EULRSiteLevelWeatherData"
xlsx_folder = r"C:\Users\Christian\RTF\EULR\EULRSiteLevelWeatherDataXLSXtest"

for filename in os.listdir(xlsx_folder):
    if filename.endswith(".xlsx"):  # Check if the file is a CSV file
        file_path = os.path.join(xlsx_folder, filename)
        
        # Read the CSV file into a DataFrame using pandas
        #df = pd.read_csv(csv_file_path)
        # Define the path for the corresponding XLSX file
        #xlsx_file_path = os.path.join(xlsx_folder, filename[:-4] + ".xlsx")
        # Save the DataFrame as an XLSX file
        #df.to_excel(xlsx_file_path, index=False)

        wb = openpyxl.load_workbook(file_path)
        sheet = wb["Sheet1"]

        column_letter = "G"
        last_data_row = None

    for row in range(sheet.max_row, 0, -1):
        cell = sheet[column_letter + str(row)]
        
        if cell.value is not None:
            last_data_row = row
            break
    
    row = 0
    for row in sheet.iter_rows(min_row=2, max_row=last_data_row, min_col=2, max_col=2):
        for cell in row:
            cell.formula = '=MONTH(J{0})'.format(cell.row)

        # num_columns_to_insert = 6
        # i = 0
        # sheet.insert_cols(i, amount=num_columns_to_insert)
        # header_values = ["hod_local", "month", "hour", "day", "year", "Timestamp"]  # Replace with your header values
        # for i, header in enumerate(header_values):
        #     sheet.cell(row=1, column=i + 1, value=header)
        
        # for row in sheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=1):
        #     for cell in row:
        #         cell.formula = '=SUM(B{0}:C{0})'.format(cell.row)
        
    wb.save(file_path)
    wb.close()